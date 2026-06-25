"""
generate_report.py — libro de trading unificado (shadow + bot real).

Hojas del Excel:
  1. Dashboard         — KPIs globales: shadow vs real, bankroll, ROI, racha
  2. Rendimiento       — Análisis completo de trader: Sharpe, drawdown, profit factor por estrategia
  3. Live_Operaciones  — historial de operaciones reales con dinero real
  4. Live_Bankroll     — movimientos de bankroll real
  5. Shadow_Ops        — historial del shadow bot (simulado)
  6. Shadow_Estrategias— IC, hit rate y P&L por estrategia
  7. Abiertas          — posiciones abiertas shadow (y live si aplica)

Ficheros de entrada:
  data/live/trades.csv
  data/live/bankroll.csv
  data/shadow/results.csv
  data/shadow/strategy_accuracy.csv
  data/shadow/performance.csv

Salida: data/shadow/informe_bot.xlsx
"""
import csv
import glob
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
    raise

DIR_SHADOW   = Path("data/shadow")
DIR_LIVE     = Path("data/live")
DIR_LIVE.mkdir(parents=True, exist_ok=True)

LIVE_TRADES_CSV    = DIR_LIVE  / "trades.csv"
LIVE_BANKROLL_CSV  = DIR_LIVE  / "bankroll.csv"
SHADOW_RESULTS     = DIR_SHADOW / "results.csv"
SHADOW_ACCURACY    = DIR_SHADOW / "strategy_accuracy.csv"
SHADOW_PERFORMANCE = DIR_SHADOW / "performance.csv"
OUTPUT_XLS         = DIR_SHADOW / "informe_bot.xlsx"

LIVE_TRADES_COLS = [
    "timestamp_utc","market_id","question","end_date",
    "direction","stake_eur","entry_price","conviction_score",
    "kelly_recomendado","estrategias","status","close_timestamp",
    "exit_price","outcome_real","fee_eur","pnl_bruto_eur",
    "pnl_neto_eur","notas",
]
LIVE_BANKROLL_COLS = ["fecha","tipo","importe_eur","balance_eur","notas"]

DEPOSITO_TOTAL          = 30.0   # depósito real en Polymarket
CAPITAL_OPERATIVO       = 20.0   # capital que puede usar el bot
RESERVA                 = 10.0   # colchón intocable
APUESTA_SHADOW          = 0.90   # apuesta legacy (no usada en ROI real)
BANKROLL_INICIAL_SHADOW = CAPITAL_OPERATIVO

_F = lambda hex: PatternFill("solid", fgColor=hex)
FILL = {
    "header":   _F("1A237E"),
    "subhead":  _F("283593"),
    "verde":    _F("C8E6C9"),
    "rojo":     _F("FFCDD2"),
    "azul":     _F("BBDEFB"),
    "amarillo": _F("FFF9C4"),
    "gris":     _F("ECEFF1"),
    "gris2":    _F("F5F5F5"),
    "naranja":  _F("FFE0B2"),
    "blanco":   _F("FFFFFF"),
    "live_ok":  _F("A5D6A7"),
    "live_ko":  _F("EF9A9A"),
    "morado":   _F("E8EAF6"),
}
_FONT = lambda bold, color, size=10: Font(bold=bold, color=color, name="Calibri", size=size)
FONT = {
    "header":  _FONT(True,  "FFFFFF", 10),
    "titulo":  _FONT(True,  "1A237E", 14),
    "subtit":  _FONT(True,  "1565C0", 11),
    "label":   _FONT(True,  "37474F", 10),
    "normal":  _FONT(False, "212121", 10),
    "verde":   _FONT(True,  "1B5E20", 10),
    "rojo":    _FONT(True,  "B71C1C", 10),
    "naranja": _FONT(True,  "E65100", 10),
    "gris":    _FONT(False, "607D8B", 9),
    "bold":    _FONT(True,  "212121", 10),
    "morado":  _FONT(True,  "283593", 10),
}
BORDE = Border(
    left=Side(style="thin",   color="CFD8DC"),
    right=Side(style="thin",  color="CFD8DC"),
    top=Side(style="thin",    color="CFD8DC"),
    bottom=Side(style="thin", color="CFD8DC"),
)
AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left",   vertical="center")
AR = Alignment(horizontal="right",  vertical="center")


def cell_set(cell, value=None, fill=None, font=None, align=None, fmt=None, border=True):
    if value is not None: cell.value = value
    if fill:   cell.fill  = fill
    if font:   cell.font  = font
    if align:  cell.alignment = align
    if border: cell.border = BORDE
    if fmt:    cell.number_format = fmt


def fila_header(ws, fila, columnas):
    for c, txt in enumerate(columnas, 1):
        cell_set(ws.cell(row=fila, column=c, value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[fila].height = 22


def kpi(ws, fila, col_lbl, col_val, label, valor, fill_v, font_v):
    cl = ws.cell(row=fila, column=col_lbl, value=label)
    cell_set(cl, fill=FILL["gris"], font=FONT["label"], align=AL)
    cv = ws.cell(row=fila, column=col_val, value=valor)
    cell_set(cv, fill=fill_v, font=font_v, align=AC)
    ws.row_dimensions[fila].height = 22


def init_live_csvs():
    if not LIVE_TRADES_CSV.exists():
        with open(LIVE_TRADES_CSV, "w", newline="", encoding="utf-8") as f:
            csv.DictWriter(f, fieldnames=LIVE_TRADES_COLS).writeheader()
        print("  Creado data/live/trades.csv (vacío)")
    if not LIVE_BANKROLL_CSV.exists():
        with open(LIVE_BANKROLL_CSV, "w", newline="", encoding="utf-8") as f:
            csv.DictWriter(f, fieldnames=LIVE_BANKROLL_COLS).writeheader()
        print("  Creado data/live/bankroll.csv (vacío)")


def cargar_csv(path):
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def cargar_strategy_params():
    path = DIR_SHADOW / "strategy_params.json"
    if not path.exists():
        return {}
    try:
        import json
        with open(path, encoding="utf-8") as f:
            return json.load(f).get("estrategias", {})
    except Exception:
        return {}


def _stake_real(r):
    """Estima el stake real de una operación a partir del pnl_bruto y precio."""
    try:
        pnl_b  = float(r.get("pnl_bruto", 0))
        acierto = int(r.get("acierto", 0))
        price   = float(r.get("precio_yes_mercado", 0.5))
        dec     = r.get("decision", "")
        if acierto == 0:
            return abs(pnl_b)                        # pérdida: pnl_b = -stake
        eff = price if dec == "BUY_YES" else max(0.001, 1 - price)
        if 0 < eff < 1:
            return pnl_b / (1.0 / eff - 1.0)       # ganancia: pnl_b = stake*(1/p-1)
    except Exception:
        pass
    return APUESTA_SHADOW


def stats_por_subtype(resultados, strategy_params):
    """Calcula stats completas por estrategia+subtype desde results.csv."""
    from collections import defaultdict
    by = defaultdict(lambda: {"n": 0, "w": 0, "pnl": 0.0, "staked": 0.0,
                               "wins": [], "losses": []})
    for r in resultados:
        strat   = r.get("strategy", "")
        subtype = r.get("subtype", "")
        # Clave más específica disponible: UPDOWN_GBM#BTC#15min
        if subtype:
            key = f"{strat}#{subtype}"
        else:
            key = strat
        try:
            pnl = float(r.get("pnl_neto", 0))
            acierto = int(r.get("acierto", 0))
            stake = _stake_real(r)
        except Exception:
            continue
        d = by[key]
        d["n"] += 1
        d["w"] += acierto
        d["pnl"] += pnl
        d["staked"] += stake
        if acierto:
            d["wins"].append(pnl)
        else:
            d["losses"].append(pnl)

    rows = []
    for key, d in sorted(by.items(), key=lambda x: x[1]["pnl"], reverse=True):
        n = d["n"]; w = d["w"]
        hit  = w / n if n else 0
        ic   = ((w + 1) / (n + 2) - 0.5) * min(1.0, n / 20)
        pnl  = d["pnl"]
        avg_win  = sum(d["wins"])  / len(d["wins"])  if d["wins"]  else 0
        avg_loss = sum(d["losses"])/ len(d["losses"]) if d["losses"] else 0
        exp  = hit * avg_win + (1 - hit) * avg_loss
        pf   = sum(d["wins"]) / abs(sum(d["losses"])) if d["losses"] else 99.0
        sp   = strategy_params.get(key, {})
        rows.append({
            "key":     key,
            "n":       n,
            "w":       w,
            "hit":     hit,
            "ic":      ic,
            "pnl":     pnl,
            "pnl_op":  pnl / n if n else 0,
            "exp":     exp,
            "pf":      pf,
            "staked":  d["staked"],
            "activa":  sp.get("activa", True),
            "kelly":   sp.get("apuesta_kelly", 0.5),
            "filtros": len(sp.get("filtros_causales", [])),
        })
    return rows


def cargar_shadow_abiertas():
    ya_resueltas = set()
    for r in cargar_csv(SHADOW_RESULTS):
        ya_resueltas.add((r.get("prediction_timestamp",""),
                          r.get("strategy",""), r.get("market_id","")))
    archivos = sorted(glob.glob(str(DIR_SHADOW / "predictions_*.csv")))[-3:]
    abiertas, vistos = [], set()
    for arch in reversed(archivos):
        with open(arch, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("decision","") not in ("BUY_YES","BUY_NO"):
                    continue
                clave = (row.get("timestamp_utc",""), row.get("strategy",""), row.get("market_id",""))
                if clave in ya_resueltas:
                    continue
                key2 = (row.get("strategy",""), row.get("market_id",""))
                if key2 in vistos:
                    continue
                vistos.add(key2)
                abiertas.append(row)
    return abiertas


def stats_live(trades):
    cerradas = [t for t in trades if t.get("status","") == "CLOSED"]
    abiertas = [t for t in trades if t.get("status","") == "OPEN"]
    n    = len(cerradas)
    n_ok = sum(1 for t in cerradas if float(t.get("pnl_neto_eur",0) or 0) > 0)
    pnl  = sum(float(t.get("pnl_neto_eur",0) or 0) for t in cerradas)
    stake = sum(float(t.get("stake_eur",0) or 0) for t in cerradas)
    roi  = pnl / stake if stake else 0
    return {"n": n, "n_ok": n_ok, "pnl": pnl, "roi": roi,
            "abiertas": abiertas, "cerradas": cerradas}


def bankroll_actual(bankroll_rows):
    if not bankroll_rows:
        return None
    try:
        return float(bankroll_rows[-1].get("balance_eur", 0))
    except Exception:
        return None


def pnl_por_dia_live(trades_cerradas):
    dias = {}
    for t in trades_cerradas:
        fecha = (t.get("close_timestamp","") or t.get("timestamp_utc",""))[:10]
        if not fecha:
            continue
        try: pnl = float(t.get("pnl_neto_eur",0) or 0)
        except: pnl = 0
        dias[fecha] = dias.get(fecha, 0) + pnl
    return dict(sorted(dias.items()))


def pnl_acum_shadow(resultados):
    acum, serie = BANKROLL_INICIAL_SHADOW, []
    for r in resultados:
        try: pnl = float(r.get("pnl_neto", 0))
        except: pnl = 0
        acum += pnl
        serie.append((r.get("resolution_timestamp","")[:10], pnl, acum))
    return serie


def calcular_max_drawdown(serie_acum: list) -> float:
    if len(serie_acum) < 2:
        return 0.0
    max_val = serie_acum[0]
    max_dd  = 0.0
    for v in serie_acum:
        max_val = max(max_val, v)
        max_dd  = max(max_dd, max_val - v)
    return max_dd


def calcular_sharpe(pnls: list) -> float:
    if len(pnls) < 2:
        return 0.0
    media = sum(pnls) / len(pnls)
    std   = (sum((p - media) ** 2 for p in pnls) / len(pnls)) ** 0.5
    return media / std if std > 0 else 0.0


# ── HOJA 1: Dashboard ──────────────────────────────────────────────────────────
def hoja_dashboard(wb, shadow_res, shadow_acc, live_trades, live_bankroll, shadow_abiertas):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    for col, w in [("A",30),("B",20),("C",4),("D",30),("E",20),("F",2),("G",44)]:
        ws.column_dimensions[col].width = w

    ts = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "BOT POLYMARKET — PANEL DE CONTROL UNIFICADO"
    c.fill      = FILL["header"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=15)
    c.alignment = AC
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = f"Actualizado: {ts}"
    c.font      = FONT["gris"]
    c.alignment = AC
    ws.row_dimensions[2].height = 14

    ws.merge_cells("A4:B4")
    c = ws["A4"]
    c.value     = "SHADOW (simulado)"
    c.fill      = FILL["subhead"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC
    ws.row_dimensions[4].height = 24

    ws.merge_cells("D4:E4")
    c = ws["D4"]
    c.value     = "LIVE (dinero real)"
    c.fill      = PatternFill("solid", fgColor="1B5E20")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC

    n_sh    = len(shadow_res)
    n_ok_sh = sum(1 for r in shadow_res if int(r.get("acierto", 0)))
    hit_sh  = n_ok_sh / n_sh if n_sh else 0
    pnl_sh  = sum(float(r.get("pnl_neto", 0)) for r in shadow_res)
    total_staked = sum(_stake_real(r) for r in shadow_res)
    roi_sh  = pnl_sh / total_staked if total_staked > 0 else 0
    bkr_sh  = BANKROLL_INICIAL_SHADOW + pnl_sh
    roi_deposito = pnl_sh / DEPOSITO_TOTAL  # ROI sobre el depósito total

    racha_sh, tipo_sh = 0, ""
    for r in reversed(shadow_res):
        ac = int(r.get("acierto", 0))
        if racha_sh == 0:
            tipo_sh = "✅" if ac else "❌"; racha_sh = 1
        elif (tipo_sh == "✅" and ac) or (tipo_sh == "❌" and not ac):
            racha_sh += 1
        else:
            break

    # Bloque capital — encabezado propio
    ws.merge_cells("A4:B4")
    c = ws["A4"]
    c.value     = "CAPITAL"
    c.fill      = PatternFill("solid", fgColor="37474F")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC
    ws.row_dimensions[4].height = 24

    for idx, (lbl, val, fv, fn) in enumerate([
        ("Depósito total",    f"{DEPOSITO_TOTAL:.0f} €",    FILL["gris"],    FONT["bold"]),
        ("Capital operativo", f"{CAPITAL_OPERATIVO:.0f} €", FILL["azul"],    FONT["bold"]),
        ("Reserva intocable", f"{RESERVA:.0f} €",           FILL["amarillo"],FONT["bold"]),
    ]):
        kpi(ws, 5 + idx, 1, 2, lbl, val, fv, fn)

    ws.merge_cells("A8:B8")
    c = ws["A8"]
    c.value     = "SHADOW (simulado)"
    c.fill      = FILL["subhead"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC
    ws.row_dimensions[8].height = 24

    f = 9
    for lbl, val, fv, fn in [
        ("Operaciones resueltas", n_sh,                           FILL["gris"],    FONT["bold"]),
        ("Posiciones abiertas",   len(shadow_abiertas),           FILL["amarillo"],FONT["bold"]),
        ("Aciertos / Fallos",     f"{n_ok_sh} / {n_sh-n_ok_sh}", FILL["verde"] if hit_sh >= 0.5 else FILL["rojo"], FONT["verde"] if hit_sh >= 0.5 else FONT["rojo"]),
        ("Hit Rate",              f"{hit_sh*100:.1f}%",           FILL["verde"] if hit_sh >= 0.5 else FILL["rojo"], FONT["verde"] if hit_sh >= 0.5 else FONT["rojo"]),
        ("P&L Neto",              f"{pnl_sh:+.2f} €",            FILL["verde"] if pnl_sh >= 0 else FILL["rojo"],   FONT["verde"] if pnl_sh >= 0 else FONT["rojo"]),
        ("ROI s/ capital operativo", f"{roi_sh*100:+.1f}%",      FILL["verde"] if roi_sh >= 0 else FILL["rojo"],   FONT["verde"] if roi_sh >= 0 else FONT["rojo"]),
        ("ROI s/ depósito total", f"{roi_deposito*100:+.1f}%",   FILL["verde"] if roi_deposito >= 0 else FILL["rojo"], FONT["verde"] if roi_deposito >= 0 else FONT["rojo"]),
        ("Bankroll simulado",     f"{bkr_sh:.2f} €",              FILL["azul"],    FONT["bold"]),
        ("Racha actual",          f"{tipo_sh} × {racha_sh}",      FILL["amarillo"],FONT["bold"]),
    ]:
        kpi(ws, f, 1, 2, lbl, val, fv, fn)
        f += 1

    lv       = stats_live(live_trades)
    bkr_live = bankroll_actual(live_bankroll)
    hit_lv   = lv["n_ok"] / lv["n"] if lv["n"] else 0
    sin_datos = lv["n"] == 0

    racha_lv, tipo_lv = 0, ""
    for t in reversed(lv["cerradas"]):
        ok = float(t.get("pnl_neto_eur", 0) or 0) > 0
        if racha_lv == 0:
            tipo_lv = "✅" if ok else "❌"; racha_lv = 1
        elif (tipo_lv == "✅" and ok) or (tipo_lv == "❌" and not ok):
            racha_lv += 1
        else:
            break

    ws.merge_cells("D4:E4")
    c = ws["D4"]
    c.value     = "LIVE (dinero real)"
    c.fill      = PatternFill("solid", fgColor="1B5E20")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC

    f = 5
    for lbl, val, fv, fn in [
        ("Depósito total",    f"{DEPOSITO_TOTAL:.0f} €",    FILL["gris"],    FONT["bold"]),
        ("Capital operativo", f"{CAPITAL_OPERATIVO:.0f} €", FILL["azul"],    FONT["bold"]),
        ("Reserva intocable", f"{RESERVA:.0f} €",           FILL["amarillo"],FONT["bold"]),
    ]:
        kpi(ws, f, 4, 5, lbl, val, fv, fn)
        f += 1

    ws.merge_cells("D8:E8")
    c = ws["D8"]
    c.value     = "LIVE — Operaciones"
    c.fill      = PatternFill("solid", fgColor="1B5E20")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = AC

    f = 9
    for lbl, val, fv, fn in [
        ("Operaciones cerradas",   lv["n"] if not sin_datos else "— (bot no activo aún)", FILL["gris"], FONT["bold"]),
        ("Posiciones abiertas",    len(lv["abiertas"]) if not sin_datos else "—", FILL["amarillo"], FONT["bold"]),
        ("Aciertos / Fallos",      f"{lv['n_ok']} / {lv['n']-lv['n_ok']}" if not sin_datos else "—", FILL["verde"] if not sin_datos and hit_lv >= 0.5 else (FILL["rojo"] if not sin_datos else FILL["gris"]), FONT["verde"] if not sin_datos and hit_lv >= 0.5 else (FONT["rojo"] if not sin_datos else FONT["gris"])),
        ("Hit Rate",               f"{hit_lv*100:.1f}%" if not sin_datos else "—", FILL["verde"] if not sin_datos and hit_lv >= 0.5 else (FILL["rojo"] if not sin_datos else FILL["gris"]), FONT["verde"] if not sin_datos and hit_lv >= 0.5 else (FONT["rojo"] if not sin_datos else FONT["gris"])),
        ("P&L Neto",               f"{lv['pnl']:+.2f} €" if not sin_datos else "—", FILL["verde"] if not sin_datos and lv["pnl"] >= 0 else (FILL["rojo"] if not sin_datos else FILL["gris"]), FONT["verde"] if not sin_datos and lv["pnl"] >= 0 else (FONT["rojo"] if not sin_datos else FONT["gris"])),
        ("ROI s/ capital operativo", f"{lv['roi']*100:+.1f}%" if not sin_datos else "—", FILL["verde"] if not sin_datos and lv["roi"] >= 0 else (FILL["rojo"] if not sin_datos else FILL["gris"]), FONT["verde"] if not sin_datos and lv["roi"] >= 0 else (FONT["rojo"] if not sin_datos else FONT["gris"])),
        ("Bankroll actual (real)", f"{bkr_live:.2f} €" if bkr_live else f"{CAPITAL_OPERATIVO:.0f} € (sin trades aún)", FILL["live_ok"] if bkr_live and not sin_datos else FILL["gris"], FONT["verde"] if bkr_live and not sin_datos else FONT["gris"]),
        ("Racha actual",           f"{tipo_lv} × {racha_lv}" if not sin_datos else "—", FILL["amarillo"], FONT["bold"]),
    ]:
        kpi(ws, f, 4, 5, lbl, val, fv, fn)
        f += 1

    if sin_datos:
        fila_nota = f + 1
        ws.merge_cells(f"D{fila_nota}:G{fila_nota}")
        c = ws.cell(row=fila_nota, column=4,
                    value="💡 Cuando el bot opere en real, añade cada trade en data/live/trades.csv")
        c.font      = Font(italic=True, color="1565C0", name="Calibri", size=9)
        c.alignment = AL

    serie_sh = pnl_acum_shadow(shadow_res)
    if len(serie_sh) >= 2:
        fila_g = 20
        ws.cell(row=fila_g, column=1, value="#")
        ws.cell(row=fila_g, column=2, value="Bankroll Shadow €")
        for i, (fecha, pnl, acum) in enumerate(serie_sh, 1):
            ws.cell(row=fila_g + i, column=1, value=i)
            ws.cell(row=fila_g + i, column=2, value=round(acum, 4))

        chart = LineChart()
        chart.title  = "Evolución Bankroll Shadow (€)"
        chart.style  = 10
        chart.height = 9
        chart.width  = 20
        chart.y_axis.title = "€"
        data_ref = Reference(ws, min_col=2, min_row=fila_g, max_row=fila_g + len(serie_sh))
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "1565C0"
        chart.series[0].graphicalProperties.line.width = 18000
        ws.add_chart(chart, "G4")


# ── HOJA 2: Rendimiento ────────────────────────────────────────────────────────
def hoja_rendimiento(wb, performance, shadow_res):
    ws = wb.create_sheet("Rendimiento", 1)
    ws.sheet_view.showGridLines = False

    anchos = [("A",28),("B",7),("C",7),("D",7),("E",9),
              ("F",10),("G",11),("H",11),("I",11),("J",13),
              ("K",10),("L",10),("M",10),("N",10),("O",22)]
    for col, w in anchos:
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value     = "ANÁLISIS DE RENDIMIENTO — VISIÓN DE TRADER / INVERSOR"
    c.fill      = FILL["header"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    c.alignment = AC
    ws.row_dimensions[1].height = 28

    n_total  = len(shadow_res)
    n_ok     = sum(1 for r in shadow_res if int(r.get("acierto", 0)))
    hit_g    = n_ok / n_total if n_total else 0

    pnls      = [float(r.get("pnl_neto", 0)) for r in shadow_res]
    pnl_total = sum(pnls)
    ganancias = [p for p in pnls if p > 0]
    perdidas  = [p for p in pnls if p < 0]
    avg_win   = sum(ganancias) / len(ganancias) if ganancias else 0
    avg_loss  = sum(perdidas)  / len(perdidas)  if perdidas  else 0
    expectancy_g  = hit_g * avg_win + (1 - hit_g) * avg_loss
    pf_g          = sum(ganancias) / abs(sum(perdidas)) if perdidas else 99.0

    acum = BANKROLL_INICIAL_SHADOW
    serie_acum = [acum]
    for p in pnls:
        acum += p
        serie_acum.append(acum)
    max_dd   = calcular_max_drawdown(serie_acum)
    sharpe   = calcular_sharpe(pnls)
    recovery = pnl_total / max_dd if max_dd > 0 else 0.0

    ws.row_dimensions[2].height = 6

    ws.merge_cells("A3:O3")
    c = ws["A3"]
    c.value     = "MÉTRICAS GLOBALES DEL SISTEMA"
    c.fill      = FILL["subhead"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = AC
    ws.row_dimensions[3].height = 20

    metricas = [
        ("Operaciones resueltas",  str(n_total),            FILL["gris"],    FONT["bold"]),
        ("Hit Rate global",        f"{hit_g*100:.1f}%",      FILL["verde"] if hit_g >= 0.5 else FILL["rojo"], FONT["verde"] if hit_g >= 0.5 else FONT["rojo"]),
        ("P&L Total",              f"{pnl_total:+.2f} €",    FILL["verde"] if pnl_total >= 0 else FILL["rojo"], FONT["verde"] if pnl_total >= 0 else FONT["rojo"]),
        ("Expectancy / op.",       f"{expectancy_g:+.4f} €", FILL["verde"] if expectancy_g > 0 else FILL["rojo"], FONT["verde"] if expectancy_g > 0 else FONT["rojo"]),
        ("Profit Factor",          f"{pf_g:.2f}" if pf_g < 99 else "∞", FILL["verde"] if pf_g >= 1.5 else (FILL["amarillo"] if pf_g >= 1 else FILL["rojo"]), FONT["bold"]),
        ("Avg. Ganancia / op.",    f"{avg_win:+.4f} €",      FILL["verde"],   FONT["verde"]),
        ("Avg. Pérdida / op.",     f"{avg_loss:+.4f} €",     FILL["rojo"],    FONT["rojo"]),
        ("Max Drawdown",           f"{max_dd:.2f} €",        FILL["naranja"] if max_dd > 0 else FILL["verde"], FONT["naranja"] if max_dd > 0 else FONT["verde"]),
        ("Sharpe (por operación)", f"{sharpe:+.3f}",         FILL["verde"] if sharpe > 0.5 else (FILL["amarillo"] if sharpe > 0 else FILL["rojo"]), FONT["bold"]),
        ("Recovery Factor",        f"{recovery:.2f}x" if recovery else "—", FILL["gris"], FONT["bold"]),
    ]

    for i, (lbl, val, fv, fn) in enumerate(metricas):
        row = 4 + (i % 5)
        col_l = 1 + (i // 5) * 4
        col_v = col_l + 1
        cl = ws.cell(row=row, column=col_l, value=lbl)
        cell_set(cl, fill=FILL["gris"], font=FONT["label"], align=AL)
        cv = ws.cell(row=row, column=col_v, value=val)
        cell_set(cv, fill=fv, font=fn, align=AC)
        ws.row_dimensions[row].height = 20

    fila_sep = 10
    ws.row_dimensions[fila_sep].height = 8
    ws.merge_cells(f"A{fila_sep}:O{fila_sep}")
    c = ws[f"A{fila_sep}"]
    c.value     = "RANKING DE ESTRATEGIAS (ordenado por P&L Total)"
    c.fill      = FILL["subhead"]
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = AC
    ws.row_dimensions[fila_sep].height = 20

    headers = [
        "Estrategia", "Ops", "Aciert.", "Hit %", "IC Bayes",
        "P&L Total €", "P&L Medio €", "Expectancy", "Profit Factor",
        "Mejor Racha", "Peor Racha", "Edge Pred.", "Edge Real",
        "Kelly Opt.", "Causa Pérdida Principal",
    ]
    fila_h = fila_sep + 1
    for c_idx, txt in enumerate(headers, 1):
        cell_set(ws.cell(row=fila_h, column=c_idx, value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[fila_h].height = 22

    if not performance:
        ws.merge_cells(f"A{fila_h+1}:O{fila_h+1}")
        c = ws.cell(row=fila_h + 1, column=1,
                    value="Sin datos de rendimiento todavía — esperando primeras resoluciones.")
        c.font      = Font(italic=True, color="607D8B", name="Calibri")
        c.alignment = AL
        return

    for i, p in enumerate(performance, 1):
        fila = fila_h + i
        try: hit = float(p.get("hit_rate", 0))
        except: hit = 0.0
        try: pnl = float(p.get("pnl_total", 0))
        except: pnl = 0.0
        try: exp = float(p.get("expectancy", 0))
        except: exp = 0.0
        try: pf  = float(p.get("profit_factor", 0))
        except: pf = 0.0
        try: ic  = float(p.get("ic_bayes", 0))
        except: ic = 0.0
        try: n_total_p   = int(p.get("n_total", 0))
        except: n_total_p = 0
        try: n_aciertos = int(p.get("n_aciertos", 0))
        except: n_aciertos = 0
        try: pnl_medio = float(p.get("pnl_medio", 0))
        except: pnl_medio = 0.0
        try: mejor_racha = int(p.get("mejor_racha", 0))
        except: mejor_racha = 0
        try: peor_racha = int(p.get("peor_racha", 0))
        except: peor_racha = 0
        try: edge_medio_pred = float(p.get("edge_medio_pred", 0))
        except: edge_medio_pred = 0.0
        try: edge_real = float(p.get("edge_real", 0))
        except: edge_real = 0.0
        try: kelly_optimo = float(p.get("kelly_optimo", 0))
        except: kelly_optimo = 0.0

        fv_hit = FILL["verde"] if hit >= 0.5 else FILL["rojo"]
        fn_hit = FONT["verde"] if hit >= 0.5 else FONT["rojo"]
        fv_pnl = FILL["verde"] if pnl >= 0 else FILL["rojo"]
        fn_pnl = FONT["verde"] if pnl >= 0 else FONT["rojo"]
        fv_exp = FILL["verde"] if exp > 0 else FILL["rojo"]
        fn_exp = FONT["verde"] if exp > 0 else FONT["rojo"]
        fv_pf  = FILL["verde"] if pf >= 1.5 else (FILL["amarillo"] if pf >= 1.0 else FILL["rojo"])
        fv_ic  = FILL["verde"] if ic >= 0.05 else (FILL["amarillo"] if ic >= 0 else FILL["rojo"])
        fn_ic  = FONT["verde"] if ic >= 0.05 else (FONT["bold"] if ic >= 0 else FONT["rojo"])
        fr_b   = FILL["gris"] if i % 2 == 0 else FILL["gris2"]

        vals_fmts_fills = [
            (p.get("strategy",""),        None,              fr_b,         FONT["bold"]),
            (n_total_p,                   None,              fr_b,         FONT["normal"]),
            (n_aciertos,                  None,              fr_b,         FONT["normal"]),
            (hit,                         "0.0%",            fv_hit,       fn_hit),
            (ic,                          "+0.000",          fv_ic,        fn_ic),
            (pnl,                         '#,##0.00 "€"',    fv_pnl,       fn_pnl),
            (pnl_medio,                   '#,##0.00 "€"',    fv_pnl,       fn_pnl),
            (exp,                         '#,##0.0000 "€"',  fv_exp,       fn_exp),
            (pf if pf < 99 else "∞",      "0.00",            fv_pf,        FONT["bold"]),
            (mejor_racha,                 None,              FILL["verde"], FONT["verde"]),
            (peor_racha,                  None,              FILL["rojo"],  FONT["rojo"]),
            (edge_medio_pred,             "+0.000",          fr_b,         FONT["normal"]),
            (edge_real,                   "+0.000",          fr_b,         FONT["morado"] if edge_real > 0 else FONT["rojo"]),
            (kelly_optimo,                "0.0%",            FILL["azul"], FONT["bold"]),
            (p.get("causa_perdida_principal") or "—", None,  fr_b,         FONT["gris"]),
        ]

        for c_idx, (val, fmt, fv, fn) in enumerate(vals_fmts_fills, 1):
            cell_set(ws.cell(row=fila, column=c_idx, value=val),
                     fill=fv, font=fn,
                     align=AL if c_idx in (1, 15) else AC, fmt=fmt)
        ws.row_dimensions[fila].height = 18

    fila_nota = fila_h + len(performance) + 2
    ws.merge_cells(f"A{fila_nota}:O{fila_nota}")
    c = ws.cell(row=fila_nota, column=1,
                value="💡  Profit Factor >1.5 = buena estrategia  |  Expectancy >0 = edge real  "
                      "|  Kelly Opt. = fracción óptima de bankroll a arriesgar por operación  "
                      "|  IC Bayes = información real que aporta la estrategia (>0.05 = útil)")
    c.font      = Font(italic=True, color="455A64", name="Calibri", size=9)
    c.alignment = AL
    ws.row_dimensions[fila_nota].height = 18


# ── HOJA 3: Live Operaciones ───────────────────────────────────────────────────
def hoja_live_ops(wb, live_trades):
    ws = wb.create_sheet("Live_Operaciones")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [
        ("#",5),("Fecha",17),("Mercado",44),("Dirección",12),
        ("Stake €",10),("Precio Entrada",14),("Conviction",11),
        ("Estado",10),("Cierre",17),("Outcome",10),
        ("Fee €",9),("P&L Bruto €",12),("P&L Neto €",12),("ROI op.",10),("Notas",25),
    ]
    for c, (txt, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell_set(ws.cell(row=1, column=c, value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[1].height = 22

    cerradas = [t for t in live_trades if t.get("status","") == "CLOSED"]
    abiertas = [t for t in live_trades if t.get("status","") == "OPEN"]
    todas    = cerradas + abiertas

    pnl_tot, stake_tot = 0.0, 0.0
    for i, t in enumerate(todas, 1):
        fila    = i + 1
        cerrada = t.get("status","") == "CLOSED"
        try: pnl   = float(t.get("pnl_neto_eur",0) or 0)
        except: pnl = 0
        try: stake = float(t.get("stake_eur",0) or 0)
        except: stake = 0
        try: fee   = float(t.get("fee_eur",0) or 0)
        except: fee = 0
        try: pnl_b = float(t.get("pnl_bruto_eur",0) or 0)
        except: pnl_b = 0
        try: conv  = float(t.get("conviction_score",0) or 0)
        except: conv = 0
        try: ep    = float(t.get("entry_price",0) or 0)
        except: ep = 0
        roi_op = pnl / stake if stake else 0

        fill_row = (FILL["live_ok"] if pnl >= 0 else FILL["live_ko"]) if cerrada else FILL["azul"]
        font_pnl = (FONT["verde"]   if pnl >= 0 else FONT["rojo"])    if cerrada else FONT["naranja"]

        if cerrada:
            pnl_tot += pnl; stake_tot += stake

        vals = [i, t.get("timestamp_utc","")[:16].replace("T"," "),
                t.get("question",""), t.get("direction",""),
                stake, ep, conv, t.get("status",""),
                (t.get("close_timestamp","") or "")[:16].replace("T"," "),
                t.get("outcome_real","") or "—",
                fee, pnl_b, pnl, roi_op, t.get("notas","")]
        fmts = [None,None,None,None,
                '#,##0.00 "€"','0.000','0.000',
                None,None,None,
                '#,##0.00 "€"','#,##0.00 "€"','#,##0.00 "€"','0.0%',None]

        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if c in (12, 13, 14):
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_row, font=font_pnl, align=AR, fmt=fmt)
            elif c in (4, 8, 10):
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_row, font=FONT["bold"], align=AC, fmt=fmt)
            else:
                cell_set(ws.cell(row=fila,column=c,value=val),
                         fill=FILL["gris"] if i%2==0 else FILL["gris2"],
                         font=FONT["normal"], align=AL if c in (3,15) else AC, fmt=fmt)
        ws.row_dimensions[fila].height = 18

    if cerradas:
        fr      = len(todas) + 2
        roi_tot = pnl_tot / stake_tot if stake_tot else 0
        fill_t  = FILL["live_ok"] if pnl_tot >= 0 else FILL["live_ko"]
        font_t  = Font(bold=True, color="1B5E20" if pnl_tot>=0 else "B71C1C", name="Calibri", size=10)
        for c, (val, fmt) in enumerate([
            ("TOTAL",None),(f"{len(cerradas)} cerradas",None),("",None),("",None),
            (stake_tot,'#,##0.00 "€"'),("",None),("",None),
            ("",None),("",None),("",None),("",None),
            (None,None),(pnl_tot,'#,##0.00 "€"'),(roi_tot,'0.0%'),("",None),
        ], 1):
            cell_set(ws.cell(row=fr,column=c,value=val),
                     fill=fill_t, font=font_t,
                     align=AR if c >= 12 else AC, fmt=fmt)
        ws.row_dimensions[fr].height = 22

    if not live_trades:
        ws.merge_cells("A2:O2")
        c = ws["A2"]
        c.value     = "⏳ Sin operaciones reales todavía. Se registrarán aquí cuando el bot esté activo."
        c.font      = Font(italic=True, color="1565C0", name="Calibri", size=10)
        c.alignment = AL


# ── HOJA 4: Live Bankroll ──────────────────────────────────────────────────────
def hoja_live_bankroll(wb, bankroll_rows, live_trades):
    ws = wb.create_sheet("Live_Bankroll")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col, w in [("A",14),("B",16),("C",16),("D",16),("E",35)]:
        ws.column_dimensions[col].width = w

    fila_header(ws, 1, ["Fecha","Tipo","Importe €","Balance €","Notas"])

    dias_pnl = pnl_por_dia_live([t for t in live_trades if t.get("status") == "CLOSED"])
    todas_filas = []
    for row in bankroll_rows:
        todas_filas.append({
            "fecha":   row.get("fecha",""),
            "tipo":    row.get("tipo",""),
            "importe": float(row.get("importe_eur",0) or 0),
            "balance": float(row.get("balance_eur",0) or 0),
            "notas":   row.get("notas",""),
        })

    fechas_manual = {r["fecha"] for r in todas_filas}
    for fecha, pnl in dias_pnl.items():
        if fecha not in fechas_manual:
            todas_filas.append({
                "fecha": fecha, "tipo": "PNL_DIA",
                "importe": pnl, "balance": 0,
                "notas": f"P&L generado automáticamente ({pnl:+.2f} €)",
            })

    todas_filas.sort(key=lambda r: r["fecha"])

    fill_tipo = {
        "DEPOSITO": FILL["live_ok"],
        "RETIRADA": FILL["live_ko"],
        "PNL_DIA":  FILL["azul"],
        "AJUSTE":   FILL["amarillo"],
    }

    for i, r in enumerate(todas_filas, 1):
        fila = i + 1
        ft   = fill_tipo.get(r["tipo"], FILL["gris"])
        vals = [r["fecha"], r["tipo"], r["importe"], r["balance"], r["notas"]]
        fmts = [None, None, '#,##0.00 "€"', '#,##0.00 "€"', None]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell_set(ws.cell(row=fila,column=c,value=val),
                     fill=ft if c in (2,3) else (FILL["gris"] if i%2==0 else FILL["gris2"]),
                     font=FONT["bold"] if c in (2,3) else FONT["normal"],
                     align=AC if c in (1,2) else (AR if c in (3,4) else AL), fmt=fmt)
        ws.row_dimensions[fila].height = 18

    if not todas_filas:
        ws.merge_cells("A2:E2")
        c = ws["A2"]
        c.value     = "💡 Añade tu depósito inicial en data/live/bankroll.csv para empezar el tracking."
        c.font      = Font(italic=True, color="1565C0", name="Calibri", size=10)
        c.alignment = AL


# ── HOJA 5: Shadow Operaciones ─────────────────────────────────────────────────
def hoja_shadow_ops(wb, resultados):
    ws = wb.create_sheet("Shadow_Ops")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [
        ("#",5),("Resolución",17),("Predicción",17),("Estrategia",26),
        ("Mercado",44),("Dirección",12),("Precio Entrada",14),
        ("Outcome",10),("Resultado",11),("P&L Bruto",12),("P&L Neto",12),("ROI op.",10),
    ]
    for c, (txt, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell_set(ws.cell(row=1,column=c,value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(resultados, 1):
        fila    = i + 1
        acierto = int(r.get("acierto",0))
        try: pnl_n = float(r.get("pnl_neto",0))
        except: pnl_n = 0
        try: pnl_b = float(r.get("pnl_bruto",0))
        except: pnl_b = 0
        try: precio = float(r.get("precio_yes_mercado",0.5))
        except: precio = 0.5
        roi_op = pnl_n / APUESTA_SHADOW

        fill_r = FILL["verde"] if acierto else FILL["rojo"]
        font_r = FONT["verde"] if acierto else FONT["rojo"]

        vals = [i,
                r.get("resolution_timestamp","")[:16].replace("T"," "),
                r.get("prediction_timestamp","")[:16].replace("T"," "),
                r.get("strategy",""), r.get("question",""), r.get("decision",""),
                precio, r.get("outcome_real",""),
                "✅ ACIERTO" if acierto else "❌ FALLO",
                pnl_b, pnl_n, roi_op]
        fmts = [None,None,None,None,None,None,
                '0.000',None,None,
                '#,##0.00 "€"','#,##0.00 "€"','0.0%']

        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if c == 9:
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_r, font=font_r, align=AC, fmt=fmt)
            elif c in (10,11,12):
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_r, font=font_r, align=AR, fmt=fmt)
            else:
                cell_set(ws.cell(row=fila,column=c,value=val),
                         fill=FILL["gris"] if i%2==0 else FILL["gris2"],
                         font=FONT["normal"],
                         align=AL if c==5 else AC, fmt=fmt)
        ws.row_dimensions[fila].height = 18

    if resultados:
        fr      = len(resultados) + 2
        pnl_tot = sum(float(r.get("pnl_neto",0)) for r in resultados)
        pnl_b_t = sum(float(r.get("pnl_bruto",0)) for r in resultados)
        roi_tot = pnl_tot / (APUESTA_SHADOW * len(resultados))
        fill_t  = FILL["verde"] if pnl_tot >= 0 else FILL["rojo"]
        font_t  = Font(bold=True, color="1B5E20" if pnl_tot>=0 else "B71C1C", name="Calibri", size=10)
        for c, (val, fmt) in enumerate([
            ("TOTAL",None),("",None),("",None),("",None),
            (f"{len(resultados)} ops",None),("",None),("",None),("",None),("",None),
            (pnl_b_t,'#,##0.00 "€"'),(pnl_tot,'#,##0.00 "€"'),(roi_tot,'0.0%'),
        ], 1):
            cell_set(ws.cell(row=fr,column=c,value=val),
                     fill=fill_t, font=font_t,
                     align=AR if c >= 10 else AC, fmt=fmt)
        ws.row_dimensions[fr].height = 22


# ── HOJA 6: Shadow Estrategias ─────────────────────────────────────────────────
def hoja_shadow_estrategias(wb, resultados, strategy_params):
    ws = wb.create_sheet("Shadow_Estrategias")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [
        ("Estrategia / Subtipo", 34), ("n",  6), ("Win%", 8),
        ("IC Bayes",  9), ("P&L Total €", 13), ("P&L / op €", 11),
        ("Profit F.", 9), ("Kelly €",  9), ("Estado", 12), ("Filtros", 8),
    ]
    for c, (txt, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell_set(ws.cell(row=1, column=c, value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[1].height = 22

    rows = stats_por_subtype(resultados, strategy_params)

    for i, d in enumerate(rows, 1):
        fila = i + 1
        activa = d["activa"]
        ic     = d["ic"]
        hit    = d["hit"]
        pnl    = d["pnl"]
        pf     = d["pf"]

        fv_ic  = FILL["verde"]   if ic  >= 0.05 else (FILL["amarillo"] if ic >= 0 else FILL["rojo"])
        fn_ic  = FONT["verde"]   if ic  >= 0.05 else (FONT["bold"]     if ic >= 0 else FONT["rojo"])
        fv_hit = FILL["verde"]   if hit >= 0.52  else (FILL["amarillo"] if hit >= 0.5 else FILL["rojo"])
        fn_hit = FONT["verde"]   if hit >= 0.52  else (FONT["bold"]     if hit >= 0.5 else FONT["rojo"])
        fv_pnl = FILL["verde"]   if pnl >= 0 else FILL["rojo"]
        fn_pnl = FONT["verde"]   if pnl >= 0 else FONT["rojo"]
        fv_pf  = FILL["verde"]   if pf  >= 1.1 else (FILL["amarillo"] if pf >= 1.0 else FILL["rojo"])
        fv_est = FILL["verde"]   if activa else FILL["rojo"]
        fn_est = FONT["verde"]   if activa else FONT["rojo"]
        fr_b   = FILL["gris"]   if i % 2 == 0 else FILL["gris2"]
        if not activa:
            fr_b = FILL["rojo"]

        estado_txt = "✅ activa" if activa else "🚫 desact."
        kelly_txt  = f"{d['kelly']:.2f} €"

        vals_fmts_fills = [
            (d["key"],          None,               fr_b,   FONT["bold"] if activa else FONT["rojo"]),
            (d["n"],            None,               fr_b,   FONT["normal"]),
            (hit,               "0.0%",             fv_hit, fn_hit),
            (ic,                "+0.000",           fv_ic,  fn_ic),
            (pnl,               '#,##0.00 "€"',     fv_pnl, fn_pnl),
            (d["pnl_op"],       '#,##0.00 "€"',     fv_pnl, fn_pnl),
            (pf if pf < 99 else "∞", "0.00",        fv_pf,  FONT["bold"]),
            (kelly_txt,         None,               FILL["azul"], FONT["bold"]),
            (estado_txt,        None,               fv_est, fn_est),
            (d["filtros"] if d["filtros"] > 0 else "—", None, fr_b, FONT["gris"]),
        ]

        for c_idx, (val, fmt, fv, fn) in enumerate(vals_fmts_fills, 1):
            cell_set(ws.cell(row=fila, column=c_idx, value=val),
                     fill=fv, font=fn,
                     align=AL if c_idx == 1 else AC, fmt=fmt)
        ws.row_dimensions[fila].height = 18

    if not rows:
        ws["A2"].value = "Sin datos todavía."
        ws["A2"].font  = Font(italic=True, color="607D8B")

    fila_nota = len(rows) + 3
    ws.merge_cells(f"A{fila_nota}:J{fila_nota}")
    c = ws.cell(row=fila_nota, column=1,
                value="IC Bayes = (w+1)/(n+2)-0.5 × conf(n/20) | >+0.05 = edge real | "
                      "Filtros = reglas causales aprendidas automáticamente por el postmortem")
    c.font = Font(italic=True, color="455A64", name="Calibri", size=9)
    c.alignment = AL


# ── HOJA 7: Abiertas ──────────────────────────────────────────────────────────
def hoja_abiertas(wb, shadow_ab, live_ab):
    ws = wb.create_sheet("Abiertas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [("Tipo",9),("Estrategia",26),("Mercado",44),("Dirección",12),
            ("Precio Entrada",14),("Edge / Conv.",12),("Horas Rest.",13),
            ("Fecha Cierre",17),("Fecha Apertura",17)]
    for c, (txt, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell_set(ws.cell(row=1,column=c,value=txt),
                 fill=FILL["header"], font=FONT["header"], align=AC)
    ws.row_dimensions[1].height = 22

    filas = []
    for t in live_ab:
        try: ep   = float(t.get("entry_price",0) or 0)
        except: ep = 0
        try: conv = float(t.get("conviction_score",0) or 0)
        except: conv = 0
        filas.append(("🟢 LIVE", t.get("estrategias","")[:25], t.get("question",""),
                      t.get("direction",""), ep, conv, "—",
                      t.get("end_date","")[:16].replace("T"," "),
                      t.get("timestamp_utc","")[:16].replace("T"," "), True))

    for r in shadow_ab:
        try: ep   = float(r.get("precio_yes_mercado",0.5) or 0.5)
        except: ep = 0.5
        try: edge = float(r.get("edge_neto",0) or 0)
        except: edge = 0
        try: horas = float(r.get("horas_a_vencimiento",0) or 0)
        except: horas = 0
        filas.append(("📊 SHADOW", r.get("strategy",""), r.get("question",""),
                      r.get("decision",""), ep, edge, f"{horas:.1f}h",
                      r.get("end_date","")[:16].replace("T"," "),
                      r.get("timestamp_utc","")[:16].replace("T"," "), False))

    for i, fila_data in enumerate(filas, 1):
        fila = i + 1
        tipo, est, q, dec, ep, ev, hr, fcierre, fapert, is_live = fila_data
        fill_tipo = FILL["live_ok"] if is_live else FILL["azul"]
        fill_dec  = FILL["verde"]   if dec in ("BUY_YES","YES") else FILL["naranja"]

        for c, (val, fmt) in enumerate([
            (tipo,None),(est,None),(q,None),(dec,None),
            (ep,'0.000'),(ev,'0.000'),(hr,None),(fcierre,None),(fapert,None)
        ], 1):
            if c == 1:
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_tipo, font=FONT["bold"], align=AC)
            elif c == 4:
                cell_set(ws.cell(row=fila,column=c,value=val), fill=fill_dec, font=FONT["bold"], align=AC)
            else:
                cell_set(ws.cell(row=fila,column=c,value=val),
                         fill=FILL["gris"] if i%2==0 else FILL["gris2"],
                         font=FONT["normal"],
                         align=AL if c in (2,3) else AC, fmt=fmt)
        ws.row_dimensions[fila].height = 18

    if not filas:
        ws["A2"].value = "No hay posiciones abiertas actualmente."
        ws["A2"].font  = Font(italic=True, color="607D8B")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    print(f"[{ts}] === Generando informe Excel unificado ===")

    init_live_csvs()

    shadow_res    = cargar_csv(SHADOW_RESULTS)
    shadow_perf   = cargar_csv(SHADOW_PERFORMANCE)
    strategy_params = cargar_strategy_params()
    shadow_ab     = cargar_shadow_abiertas()
    live_trades   = cargar_csv(LIVE_TRADES_CSV)
    live_bkr      = cargar_csv(LIVE_BANKROLL_CSV)
    live_ab       = [t for t in live_trades if t.get("status","") == "OPEN"]

    print(f"  Shadow resueltas:          {len(shadow_res)}")
    print(f"  Shadow abiertas:           {len(shadow_ab)}")
    print(f"  Performance estrategias:   {len(shadow_perf)}")
    print(f"  Strategy params cargados:  {len(strategy_params)}")
    print(f"  Live operaciones:          {len(live_trades)}")
    print(f"  Live movimientos bankroll: {len(live_bkr)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hoja_dashboard(wb, shadow_res, [], live_trades, live_bkr, shadow_ab)
    hoja_rendimiento(wb, shadow_perf, shadow_res)
    hoja_live_ops(wb, live_trades)
    hoja_live_bankroll(wb, live_bkr, live_trades)
    hoja_shadow_ops(wb, shadow_res)
    hoja_shadow_estrategias(wb, shadow_res, strategy_params)
    hoja_abiertas(wb, shadow_ab, live_ab)

    wb.save(OUTPUT_XLS)
    print(f"  Guardado: {OUTPUT_XLS}")
    print(f"[{ts}] === Fin ===")


if __name__ == "__main__":
    main()
